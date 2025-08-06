import pandas as pd
import os
import re
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import LineChart, PieChart, Reference, Series
from openpyxl.utils import get_column_letter
from openpyxl.chart.label import DataLabelList

def standardize_model_name(model_name):
    """
    Standardizes a model name by converting to lowercase, removing specific
    "ignore" words and special characters, splitting, sorting, and joining.
    """
    if pd.isna(model_name):
        return None
    
    ignore_words = ['comstore']
    
    # Convert to lowercase and replace non-alphanumeric characters (except underscore) with a space
    clean_name = re.sub(r'[^a-z0-9]', ' ', str(model_name).lower())
    
    # Split the string, remove ignore words and empty parts
    parts = [part for part in clean_name.split() if part and part not in ignore_words]
    
    # Sort the remaining parts alphabetically for consistent naming
    parts.sort()
    
    return '_'.join(parts)

def get_display_name_for_model(standardized_model):
    """
    Returns a user-friendly display name for a standardized model.
    """
    if standardized_model == 'crbx_fc4':
        return 'FC4_CRBX'
    else:
        # Convert underscores to spaces and title case
        display_name = standardized_model.replace('_', ' ').title()
        return display_name

def autofit_columns(sheet):
    """Adjusts column widths to fit content in a given worksheet."""
    for column in sheet.columns:
        max_length = 0
        column_name = get_column_letter(column[0].column)
        for cell in column:
            try:
                if cell.value is not None and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        sheet.column_dimensions[column_name].width = adjusted_width

def organize_excel_by_model_advanced():
    """
    Reads an Excel file, standardizes 'Model' entries, and creates
    new sheets for each standardized model type. It also adds charts
    and automatically adjusts column widths.
    """
    print("--- Organize Job Cards by Model (Advanced) ---")
    
    while True:
        input_file_path = input("Please enter the full path to your Excel file (e.g., jobcards_July_2025.xlsx): ")
        if os.path.exists(input_file_path) and input_file_path.endswith(('.xlsx', '.xls')):
            break
        else:
            print("Invalid file path or file does not exist. Please enter a valid .xlsx or .xls file path.")

    try:
        df = pd.read_excel(input_file_path)
        print(f"Successfully loaded data from '{input_file_path}'.")
        print(f"DEBUG: Columns found in the loaded Excel file: {df.columns.tolist()}")
    except Exception as e:
        print(f"Error reading the Excel file: {e}")
        print("Please ensure the file is not open and is a valid Excel workbook.")
        return

    # --- Initial column checks ---
    required_columns = ['Date', 'Model']
    for col in required_columns:
        if col not in df.columns:
            print(f"Error: Required column '{col}' not found in the Excel file.")
            print(f"Please ensure the column name is exactly '{col}' (case-sensitive).")
            return

    # Apply standardization to create a new column in the main DataFrame
    df['standardized_model'] = df['Model'].apply(standardize_model_name)
    unique_standardized_models = df['standardized_model'].dropna().unique()

    if not unique_standardized_models.size:
        print("No unique standardized models found in the 'Model' column after standardization.")
        pass 

    print(f"Found {len(unique_standardized_models)} unique standardized models.")

    dir_name = os.path.dirname(input_file_path)
    file_name_without_ext = os.path.splitext(os.path.basename(input_file_path))[0]
    output_file_path = os.path.join(dir_name, f"{file_name_without_ext}_organized_advanced.xlsx")

    print(f"Creating new Excel file: '{output_file_path}'...")

    try:
        # --- Prepare data for charts BEFORE opening ExcelWriter ---
        # This ensures the original structure and columns are preserved for chart data.
        df_chart_data = df.copy() 
        print(f"DEBUG: Columns in df_chart_data before chart prep: {df_chart_data.columns.tolist()}")

        # Prepare data for Line Graph (Job cards by date)
        jobcards_by_date = pd.DataFrame(columns=['Day', 'Count']) # Initialize as empty
        if 'Date' in df_chart_data.columns: 
            print(f"DEBUG: 'Date' column type before conversion for line chart: {df_chart_data['Date'].dtype}")
            try:
                # Convert 'Date' column to datetime, coercing errors to NaT
                # Then extract just the date part, and drop NaT
                date_series_for_chart = pd.to_datetime(df_chart_data['Date'], errors='coerce').dt.date.dropna()
                print(f"DEBUG: Series type after pd.to_datetime().dt.date.dropna(): {date_series_for_chart.dtype}")

                if not date_series_for_chart.empty:
                    jobcards_by_date = date_series_for_chart.value_counts().sort_index().reset_index()
                    jobcards_by_date.columns = ['Day', 'Count'] # Renaming for clarity
                    jobcards_by_date['Day'] = jobcards_by_date['Day'].apply(lambda x: f"{x.day:02d}") # Format day as '01', '02'
                else:
                    print("Warning: No valid datetime data found for Line Chart after parsing and cleaning. Chart will be empty.")
                print(f"DEBUG: jobcards_by_date DataFrame head:\n{jobcards_by_date.head()}")
            except Exception as e:
                print(f"Warning: Could not prepare data for Line Chart due to date processing error: {e}. Chart will be empty.")
                jobcards_by_date = pd.DataFrame(columns=['Day', 'Count']) # Ensure it's empty if error
        else:
            print("Warning: 'Date' column not found in df_chart_data for line chart preparation. Chart will be empty.")


        # Prepare data for Pie Chart (Job cards by standardized model)
        jobcards_by_model_standardized = df_chart_data['standardized_model'].value_counts().reset_index()
        jobcards_by_model_standardized.columns = ['Standardized_Model', 'Count']
        jobcards_by_model_standardized['Display_Model'] = jobcards_by_model_standardized['Standardized_Model'].apply(get_display_name_for_model)
        print(f"DEBUG: jobcards_by_model_standardized DataFrame head:\n{jobcards_by_model_standardized.head()}")

        # --- Now proceed with ExcelWriter to write sheets and add charts ---
        with pd.ExcelWriter(output_file_path, engine='openpyxl') as writer:
            # 1. Create a sheet for 'All Job Cards'
            df_for_all_sheet = df.drop(columns=['standardized_model']).copy() 
            df_for_all_sheet.to_excel(writer, sheet_name='All Job Cards', index=False)
            print("Original data written to 'All Job Cards' sheet.")

            # 2. Loop and create sheets for each model
            print("\nCreating individual model sheets:")
            if unique_standardized_models.size > 0:
                for standardized_model in unique_standardized_models:
                    sheet_name = get_display_name_for_model(standardized_model)
                    sheet_name = re.sub(r'[\\/*?[\]:]', '', sheet_name)
                    if len(sheet_name) > 31:
                        sheet_name = sheet_name[:31]

                    filtered_df = df[df['standardized_model'] == standardized_model].copy()
                    filtered_df.drop(columns=['standardized_model'], inplace=True) 
                    filtered_df.to_excel(writer, sheet_name=sheet_name, index=False)
                    print(f"  Created sheet for model: '{sheet_name}' with {len(filtered_df)} entries.")
            else:
                print("  No individual model sheets created as no unique standardized models were found.")
        
        # 3. Post-process the workbook for column widths and charts
        print("\nPost-processing workbook for column widths and charts...")
        wb = writer.book

        # Create new sheets for charts
        trend_sheet = wb.create_sheet(title='Job Cards Trend', index=1)
        pie_sheet = wb.create_sheet(title='Model Distribution', index=2)
        
        # Write data to the trend sheet for the line graph
        if not jobcards_by_date.empty:
            for r in dataframe_to_rows(jobcards_by_date, index=False, header=True):
                trend_sheet.append(r)
        else:
            trend_sheet.append(['Day', 'Count']) 
            print("No data available for Line Chart. Creating empty sheet.")

        # Write data to the pie sheet for the pie chart
        if not jobcards_by_model_standardized.empty:
            pie_sheet.append(['Display Model', 'Count']) 
            for index, row in jobcards_by_model_standardized.iterrows():
                pie_sheet.append([row['Display_Model'], row['Count']])
        else:
            pie_sheet.append(['Display Model', 'Count']) 
            print("No data available for Pie Chart. Creating empty sheet.")
        
        # Create and add the Line Chart
        if not jobcards_by_date.empty:
            line_chart = LineChart()
            line_chart.title = "Job Cards Trend by Day"
            line_chart.style = 13
            line_chart.x_axis.title = "Day of Month"
            line_chart.y_axis.title = "Quantity"
            
            # Data and categories for line chart
            line_data = Reference(trend_sheet, min_col=2, min_row=1, max_row=len(jobcards_by_date) + 1)
            line_categories = Reference(trend_sheet, min_col=1, min_row=2, max_row=len(jobcards_by_date) + 1)
            
            line_series = Series(line_data, title="Job Cards Count")
            line_series.smooth = False
            line_chart.series.append(line_series)
            line_chart.set_categories(line_categories)
            line_chart.width = 25
            line_chart.height = 15
            trend_sheet.add_chart(line_chart, "D2")
        else:
            print("Line Chart not generated due to no data.")
        
        # Create and add the Pie Chart
        if not jobcards_by_model_standardized.empty:
            pie_chart = PieChart()
            pie_chart.title = "Repaired Equipment Models Distribution"
            pie_chart.width = 15
            pie_chart.height = 15
            
            pie_labels = Reference(pie_sheet, min_col=1, min_row=2, max_row=len(jobcards_by_model_standardized) + 1)
            pie_data = Reference(pie_sheet, min_col=2, min_row=1, max_row=len(jobcards_by_model_standardized) + 1)
            
            pie_chart.add_data(pie_data, titles_from_data=True) 
            pie_chart.set_categories(pie_labels)
            pie_chart.dataLabels = None
            pie_chart.legend.position = 'r'
            
            pie_sheet.add_chart(pie_chart, "D2")
        else:
            print("Pie Chart not generated due to no data.")

        # Apply column auto-fit to all sheets
        for sheet in wb.worksheets:
            autofit_columns(sheet)

        wb.save(output_file_path)
        
        print(f"\nSuccessfully organized job cards and saved to '{output_file_path}'.")
        print("Column widths have been auto-adjusted, and charts have been added.")

        # --- Final summary of model counts (new log) ---
        print("\n--- Summary of Job Cards by Model ---")
        if not jobcards_by_model_standardized.empty:
            for index, row in jobcards_by_model_standardized.iterrows():
                print(f"  {row['Display_Model']}: {row['Count']} entries")
        else:
            print("  No model data to summarize.")

    except Exception as e:
        print(f"An error occurred during Excel writing: {e}")
        print("Please ensure the output file is not open and you have write permissions.")

if __name__ == "__main__":
    organize_excel_by_model_advanced()